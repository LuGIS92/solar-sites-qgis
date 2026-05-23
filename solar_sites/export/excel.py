"""Excel-Export der Analyseergebnisse."""

from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pandas as pd

from solar_sites.config import EXPORTS_DIR


def export_to_excel(analyses: list[dict], job_id: str, output_path: Path | None = None) -> Path:
    """Exportiert Analyseergebnisse als formatiertes Excel."""
    EXPORTS_DIR.mkdir(parents=True, exist_ok=True)

    if output_path is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_path = EXPORTS_DIR / f"solar_potenzial_{job_id}_{timestamp}.xlsx"

    df = _build_dataframe(analyses)

    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name="Solarpotenzial", index=False)
        _format_worksheet(writer.sheets["Solarpotenzial"], df)

    return output_path


def _build_dataframe(analyses: list[dict]) -> pd.DataFrame:
    rows = []
    for a in analyses:
        rows.append({
            "Adresse": a.get("raw_address", ""),
            "Straße": a.get("street", ""),
            "Hausnummer": a.get("house_nr", ""),
            "Stadt": a.get("city", ""),
            "PLZ": a.get("postal_code", ""),
            "Gebäude-ID": a.get("building_id", ""),
            "Gebäudetyp": a.get("building_type", ""),
            "Quelle": a.get("building_source", ""),
            "Dachfläche (m²)": round(a.get("usable_area_sqm") or 0, 1),
            "Installierbar (kWp)": round(a.get("installable_kwp") or 0, 2),
            "Jahresertrag (kWh)": round(a.get("annual_energy_kwh") or 0, 0),
            "PVGIS-Strahlung (kWh/kWp)": round(a.get("pvgis_irradiation") or 0, 1),
            "Panels erkannt": "Ja" if a.get("panels_detected") else "Nein",
            "MaStR registriert": "Ja" if a.get("mastr_registered") else "Nein",
            "MaStR Leistung (kW)": a.get("mastr_capacity_kw"),
            "DLR Jahresertrag (kWh)": round(a["dlr_annual_kwh"], 0) if a.get("dlr_annual_kwh") else "",
            "DLR Modulfläche (m²)": round(a["dlr_panel_area_sqm"], 1) if a.get("dlr_panel_area_sqm") else "",
            "Datenquelle": a.get("data_source", "pvgis"),
            "Lat": a.get("lat"),
            "Lon": a.get("lon"),
        })
    df = pd.DataFrame(rows)
    if not df.empty:
        df.sort_values("Jahresertrag (kWh)", ascending=False, inplace=True)
    return df


def _format_worksheet(ws, df: pd.DataFrame) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    header_fill = PatternFill("solid", fgColor="1F4E79")
    header_font = Font(color="FFFFFF", bold=True)

    for col_idx, _ in enumerate(df.columns, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for col_idx, col_name in enumerate(df.columns, 1):
        max_len = max(len(str(col_name)), df[col_name].astype(str).str.len().max())
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max_len + 2, 40)
